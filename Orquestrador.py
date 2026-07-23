import os
import pandas as pd
import openpyxl
from GoogleDrive import GoogleDriveManager
from GoogleSheets import GoogleSheets

def mapear_nome_tese(nome_pasta):
    nome_upper = nome_pasta.upper()
    if "CONCOMITANTE" in nome_upper:
        return "CONCOMITANTES"
    elif "322" in nome_upper:
        return "TEMA_322"
    elif "EMENDA" in nome_upper or "EC" in nome_upper:
        return "EMENDAS"
    elif "BURACO NEGRO" in nome_upper or "BN" in nome_upper:
        return "BURACO_NEGRO"
    return None

class OrquestradorDrive:
    def __init__(self, diretorio_json, map_estados_bots):
        """
        map_estados_bots: Dicionário que mapeia o nome da pasta do estado 
        para a classe/função responsável por processar os dados dele.
        Ex: {"SANTA CATARINA": BotTRF4().processar}
        """
        self.drive = GoogleDriveManager(diretorio_json)
        self.map_estados_bots = map_estados_bots
        # ID fixo da raiz conforme solicitado
        self.parent_folder_id = "12wtexdXWnh8bndlSYBcDKbpXbejpjldz"

    def executar_pipeline(self):
        print(f"Listando pastas na raiz ({self.parent_folder_id})...")
        pastas_raiz = self.drive.listar_arquivos(self.parent_folder_id, mime_type='application/vnd.google-apps.folder')
        
        for pasta_tese in pastas_raiz:
            tese_key = mapear_nome_tese(pasta_tese['name'])
            if not tese_key:
                print(f"Ignorando pasta (não é uma tese mapeada): {pasta_tese['name']}")
                continue
                
            print(f"Encontrada tese: {pasta_tese['name']} (Chave: {tese_key})")
            
            pastas_tese = self.drive.listar_arquivos(pasta_tese['id'], mime_type='application/vnd.google-apps.folder')
            pasta_estados = next((p for p in pastas_tese if p['name'].upper() == "ESTADOS"), None)
            
            if pasta_estados:
                print(f"  -> Explorando pasta ESTADOS dentro de {pasta_tese['name']}")
                pastas_estados_ufs = self.drive.listar_arquivos(pasta_estados['id'], mime_type='application/vnd.google-apps.folder')
                
                for pasta_uf in pastas_estados_ufs:
                    nome_estado = pasta_uf['name'].strip().upper()
                    
                    if nome_estado in self.map_estados_bots:
                        print(f"    -> Encontrado estado mapeado: {nome_estado}")
                        self._processar_estado(pasta_uf, self.map_estados_bots[nome_estado], tese_key)
                    else:
                        print(f"    -> Estado ignorado (sem bot mapeado): {nome_estado}")
            else:
                print(f"  -> Pasta ESTADOS não encontrada em {pasta_tese['name']}")

    def _processar_estado(self, pasta_uf, processador_bot, tese):
        finalizado_folder_id = self.drive.buscar_ou_criar_pasta("FINALIZADO", pasta_uf['id'])
        arquivos = self.drive.listar_arquivos(pasta_uf['id'])
        
        for arquivo in arquivos:
            if arquivo['mimeType'] == 'application/vnd.google-apps.folder':
                continue
                
            print(f"      -> Processando arquivo: {arquivo['name']}")
            file_stream, nome_arquivo = self.drive.baixar_arquivo(arquivo['id'], arquivo['name'], arquivo['mimeType'])
            
            is_valid, df_ou_erro = self.drive.validar_arquivo(file_stream, arquivo['name'])
            
            if not is_valid:
                print(f"         [ALERTA] Arquivo {arquivo['name']} inválido. Motivo: {df_ou_erro}")
                continue
                
            print(f"         Arquivo {arquivo['name']} validado. Iniciando Bot.")
            
            # Preparar dados abstraídos do Pandas
            df = df_ou_erro.fillna("")
            lista_dados = df.values.tolist()
            
            # Inicializa a API do Sheets se for uma planilha do Google
            is_google_sheet = (arquivo['mimeType'] == 'application/vnd.google-apps.spreadsheet')
            is_excel = arquivo['name'].endswith('.xlsx')
            
            planilha_api = None
            wb = None
            ws = None
            
            if is_google_sheet:
                # Usa aba com nome "Geral" (padrão) ou pode adaptar se a aba tiver outro nome
                planilha_api = GoogleSheets(arquivo['id'], "Geral", self.drive.diretorio_json)
            elif is_excel:
                file_stream.seek(0)
                wb = openpyxl.load_workbook(file_stream)
                ws = wb.active
            
            def callback_status(indice, status, numero_processo="", assunto=""):
                # -------------------------------------------------------------
                # 1. PASSO A: Resgatar os dados do cliente (CPF e Nome)
                # -------------------------------------------------------------
                linha_original = lista_dados[indice - 2]
                cpf = linha_original[0].strip() if len(linha_original) > 0 else ""
                nome = str(linha_original[2]).strip() if len(linha_original) > 2 and linha_original[2] is not None else ""

                # -------------------------------------------------------------
                # 2. PASSO B: Mudar a verificação de cor para termos semânticos
                # -------------------------------------------------------------
                # Se você mudar os status de string no 'trf4.py', podemos checar o termo diretamente
                # Ex: checar se a frase contém 'Tese localizada' ou 'Descartado'
                processo_encontrado = "Tese localizada" in status or "Descartado" in status

                if planilha_api:
                    try:
                        if processo_encontrado:
                            # 1. Garante que a aba "Processos Encontrados" existe no Google Sheets
                            planilha_api.criar_aba_se_nao_existir("Processos Encontrados")

                            # 2. Adiciona a linha de dados na nova aba
                            planilha_api.anexar_linha("Processos Encontrados", [nome, cpf, numero_processo, assunto, status])

                            # 3. Oculta a linha original na aba "Geral"
                            planilha_api.ocultar_linha(indice)
                        else:
                            # Se não encontrou, mantém visível e apenas atualiza o status na aba Geral
                            planilha_api.atualizar_celula(f"BT{indice}", numero_processo)
                            planilha_api.atualizar_celula(f"BU{indice}", assunto)
                            planilha_api.atualizar_celula(f"BV{indice}", status)
                            planilha_api.pintar_linha(indice, status)
                    except Exception as e:
                        print(f"         [AVISO] Erro ao processar Sheets: {e}")
                else:
                    if ws:
                        try:
                            if processo_encontrado:
                                # 1. Garante que a aba "Processos Encontrados" existe no Excel local
                                if "Processos Encontrados" not in wb.sheetnames:
                                    ws_encontrados = wb.create_sheet("Processos Encontrados")
                                    # Cabeçalhos da nova aba
                                    ws_encontrados.append(["Nome", "CPF", "Número Processo", "Assunto", "Status"])
                                else:
                                    ws_encontrados = wb["Processos Encontrados"]

                                # 2. Adiciona a linha com os dados
                                ws_encontrados.append([nome, cpf, numero_processo, assunto, status])

                                # 3. Oculta a linha na aba original
                                ws.row_dimensions[indice].hidden = True
                            else:
                                # Se não encontrou, apenas escreve as colunas no fim
                                ws[f"BT{indice}"] = numero_processo
                                ws[f"BU{indice}"] = assunto
                                ws[f"BV{indice}"] = status
                        except Exception as e:
                            print(f"         [AVISO] Erro ao processar Excel local: {e}")

                    print(f"         Status (Offline/Excel): Linha {indice} -> {status} | Proc: {numero_processo} | Assunto: {assunto}")

            # Executa o bot com Single Responsibility
            # O bot processa e não precisa saber de onde vieram os dados
            try:
                processador_bot(lista_dados, atualizar_status_callback=callback_status, tese=tese)
                print(f"         Bot finalizou o processamento de {arquivo['name']}.")
            except Exception as e:
                print(f"         [ERRO] Falha ao rodar o bot para {arquivo['name']}: {str(e)}")
                continue
                
            if wb:
                temp_file = f"temp_{arquivo['name']}"
                wb.save(temp_file)
                print(f"         Fazendo upload das alterações no Excel para o Google Drive...")
                self.drive.atualizar_arquivo_local(arquivo['id'], temp_file)
                if os.path.exists(temp_file):
                    os.remove(temp_file)

            print(f"         Movendo {arquivo['name']} para FINALIZADO...")
            self.drive.mover_arquivo(arquivo['id'], pasta_uf['id'], finalizado_folder_id)
            print("         Movido com sucesso!")
